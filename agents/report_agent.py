"""Agente generador del reporte semanal en Excel (.xlsx).

Produce el archivo siguiendo la estructura del template oficial
(shaffe_ads_reporte_semanal_TEMPLATE.xlsx), con 6 hojas de trabajo:
Resumen Ejecutivo, Publicaciones Problema, Alertas Urgentes,
Sin Ads - Candidatas, Acciones Semana, Historial Cambios."""
from __future__ import annotations

import os
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import core.campaign_rules as reglas

BASE_DIR = os.path.dirname(os.path.dirname(__file__))

# Colores de semáforo
_VERDE  = "C6EFCE"
_AMARILLO = "FFEB9C"
_ROJO   = "FFC7CE"
_GRIS   = "D9D9D9"
_AZUL_HEADER = "1F4E79"


def _header(ws, fila: int, valores: list, color: str = _AZUL_HEADER) -> None:
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF" if color == _AZUL_HEADER else "000000")
    for col, valor in enumerate(valores, start=1):
        cell = ws.cell(row=fila, column=col, value=valor)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _semaforo_fill(estado: str) -> PatternFill:
    colores = {"OK": _VERDE, "Revisar": _AMARILLO, "Problema": _ROJO}
    for k, c in colores.items():
        if k.lower() in estado.lower():
            return PatternFill("solid", fgColor=c)
    return PatternFill("solid", fgColor=_GRIS)


def _autowidth(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


class ReportAgent:
    def generar(
        self,
        grupos: dict,
        acciones: list,
        alertas_urgentes: list,
        candidatas_sin_ads: dict,
        changes_history: list,
        semana: str,
    ) -> str:
        """Genera el xlsx y devuelve la ruta del archivo creado."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._hoja_resumen(wb, grupos, acciones, semana)
        self._hoja_problema(wb, grupos, acciones, semana)
        self._hoja_urgentes(wb, alertas_urgentes, semana)
        self._hoja_sin_ads(wb, candidatas_sin_ads, semana)
        self._hoja_acciones(wb, acciones, semana)
        self._hoja_historial(wb, changes_history)

        nombre = f"shaffe_ads_reporte_{date.today().isoformat()}.xlsx"
        ruta = os.path.join(BASE_DIR, "logs", nombre)
        os.makedirs(os.path.dirname(ruta), exist_ok=True)
        wb.save(ruta)
        return ruta

    # ------------------------------------------------------------------ #
    # Hoja 1: Resumen Ejecutivo
    # ------------------------------------------------------------------ #
    def _hoja_resumen(self, wb, grupos: dict, acciones: list, semana: str) -> None:
        ws = wb.create_sheet("Resumen Ejecutivo")

        ws["A1"] = "SHAFFE - Reporte Semanal MercadoLibre Ads"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Semana: {semana}  |  Generado automáticamente por el agente"

        # KPIs generales
        ws["A4"] = "ESTADO GENERAL DE LA CUENTA"
        ws["A4"].font = Font(bold=True)
        _header(ws, 5, ["KPI", "Semana actual", "Semana anterior", "Variación", "Objetivo", "Estado", "Recomendación del agente"])

        total_costo = sum(g["metricas"].get("cost", 0) for g in grupos.values())
        total_ingresos = sum(
            g["metricas"].get("direct_amount", 0) + g["metricas"].get("indirect_amount", 0)
            for g in grupos.values()
        )
        total_clicks = sum(g["metricas"].get("clicks", 0) for g in grupos.values())
        roas_general = (total_ingresos / total_costo) if total_costo > 0 else 0.0
        acos_general = (total_costo / total_ingresos) if total_ingresos > 0 else None
        pubs_en_ads = len(grupos)

        kpis = [
            ("ROAS general", round(roas_general, 2), None, None, "6.5+",
             "✅ OK" if roas_general >= 6.5 else "⚠️ Revisar",
             "Seguir escalando campañas Oro. No frenar lo que funciona." if roas_general >= 6.5
             else "Revisá la hoja Publicaciones Problema — identificá qué arrastra el ROAS hacia abajo."),
            ("ACOS general (%)", f"{acos_general*100:.1f}%" if acos_general else "N/A", None, None, "≤18%",
             "✅ OK" if (acos_general and acos_general <= 0.18) else "⚠️ Revisar",
             "Margen de publicidad saludable." if (acos_general and acos_general <= 0.18)
             else "Bajá presupuesto en campañas con ROAS bajo objetivo para reducir el ACOS."),
            ("Inversión total ($)", round(total_costo), None, None, "-", "-", "-"),
            ("Ventas atribuidas ($)", round(total_ingresos), None, None, "-", "-", "-"),
            ("Clicks totales", total_clicks, None, None, "-", "-", "-"),
            ("Publicaciones activas en Ads", pubs_en_ads, None, None, "-", "-", "-"),
        ]
        for i, row in enumerate(kpis, start=6):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)
            if row[5] and "OK" in str(row[5]):
                ws.cell(row=i, column=6).fill = PatternFill("solid", fgColor=_VERDE)
            elif row[5] and "Revisar" in str(row[5]):
                ws.cell(row=i, column=6).fill = PatternFill("solid", fgColor=_AMARILLO)

        # Semáforo por campaña
        fila = len(kpis) + 8
        ws.cell(row=fila - 1, column=1, value="SEMÁFORO POR CAMPAÑA").font = Font(bold=True)
        _header(ws, fila, ["Campaña", "Tier", "Inversión ($)", "Ventas ($)", "ROAS", "ACOS", "Estado", "Recomendación del agente"])
        fila += 1

        campanas_config = [
            ("Oro - Ticket Alto ($33k+)",    "ORO",    "oro_alto",  7.5),
            ("Oro - Ticket Medio ($18k-$33k)", "ORO",  "oro_medio", 6.5),
            ("Oro - Ticket Bajo (< $18k)",   "ORO",    "oro_bajo",  8.0),
            ("Plata - Recuperación",          "PLATA",  "plata",     4.0),
            ("Testeo - Productos nuevos",     "TESTEO", "testeo",    3.0),
        ]

        def _recomendacion_semaforo(tier: str, roas: float, roas_obj: float) -> str:
            ratio = roas / roas_obj if roas_obj else 0
            if tier == "oro":
                if ratio >= 1:
                    return "Ganadoras. Subir presupuesto leve si hay stock ≥ 10 uds. No tocar lo que funciona."
                if ratio >= 0.8:
                    return "Cerca del objetivo. Revisá publicaciones con ROAS < 5.5 — moverlas a Plata si persisten 3 días."
                return "ROAS bajo para Oro — mover a Plata las publicaciones que no cumplen el objetivo esta semana."
            if tier == "plata":
                if ratio >= 1:
                    return "Sala de observación en buen estado. Si ROAS ≥ 6.5 sostenido 3 días → mover a Oro."
                if ratio >= 0.7:
                    return "ROAS cerca del mínimo. Monitorear. Si no mejora en 3 días → pausar publicaciones problema."
                return "ROAS en Plata por debajo del mínimo — pausar publicaciones con gasto real y sin retorno."
            # testeo
            if ratio >= 1:
                return "Testeo funcionando. Si ROAS > 3.5 sostenido 3 días → mover a Plata."
            if ratio >= 0.6:
                return "En rango testeo. Revisar fotos principales. Sin mejora en 7 días → pausar."
            return "Sin resultados claros — cambiar foto/título en las publicaciones con gasto. Si persiste, pausar."

        for nombre_display, tier_display, clave_tier, roas_obj in campanas_config:
            grupos_campana = [
                g for g in grupos.values()
                if any(t.startswith(clave_tier) for t in g["tiers_detectados"])
            ]
            costo_c = sum(g["metricas"].get("cost", 0) for g in grupos_campana)
            ing_c = sum(
                g["metricas"].get("direct_amount", 0) + g["metricas"].get("indirect_amount", 0)
                for g in grupos_campana
            )
            roas_c = (ing_c / costo_c) if costo_c > 0 else 0.0
            acos_c = (costo_c / ing_c) if ing_c > 0 else None
            estado = "✅ OK" if roas_c >= roas_obj else ("⚠️ Revisar" if roas_c >= roas_obj * 0.7 else "🔴 Problema")
            tier_key = clave_tier.split("_")[0]
            recomendacion = _recomendacion_semaforo(tier_key, roas_c, roas_obj)

            ws.cell(row=fila, column=1, value=nombre_display)
            ws.cell(row=fila, column=2, value=tier_display)
            ws.cell(row=fila, column=3, value=round(costo_c))
            ws.cell(row=fila, column=4, value=round(ing_c))
            ws.cell(row=fila, column=5, value=round(roas_c, 2))
            ws.cell(row=fila, column=6, value=f"{acos_c*100:.1f}%" if acos_c else "N/A")
            ws.cell(row=fila, column=7, value=estado)
            ws.cell(row=fila, column=8, value=recomendacion)
            ws.cell(row=fila, column=7).fill = _semaforo_fill(estado)
            ws.cell(row=fila, column=8).alignment = Alignment(wrap_text=True)
            fila += 1

        _autowidth(ws)

    # ------------------------------------------------------------------ #
    # Hoja 2: Publicaciones Problema
    # ------------------------------------------------------------------ #
    def _hoja_problema(self, wb, grupos: dict, acciones: list, semana: str) -> None:
        ws = wb.create_sheet("Publicaciones Problema")
        ws["A1"] = "PUBLICACIONES PROBLEMA - Requieren acción esta semana"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = f"Semana: {semana}  |  ROAS bajo objetivo o ACOS elevado con gasto real significativo"

        _header(ws, 4, ["MLA", "Título publicación", "Campaña", "Tier", "Inversión ($)", "Ventas ($)", "ROAS", "ACOS", "Diagnóstico", "Acción"])

        fila = 5
        alertas_dict = {
            tuple(sorted(a["item_ids"])): a
            for a in acciones if a.get("tipo") == "alerta"
        }
        for family_id, grupo in grupos.items():
            metricas = grupo["metricas"]
            roas = metricas.get("roas", 0.0)
            costo = metricas.get("cost", 0.0)
            campania = grupo["tiers_detectados"][0] if grupo["tiers_detectados"] else ""
            roas_obj = reglas.roas_target_campania(campania)

            if roas >= roas_obj or costo < 100:
                continue

            acos = metricas.get("acos")
            ingresos = metricas.get("direct_amount", 0) + metricas.get("indirect_amount", 0)
            tier_actual, _ = reglas.tier_y_ticket(campania) if campania else ("", "")

            clave = tuple(sorted(grupo["item_ids"]))
            alerta = alertas_dict.get(clave, {})
            diagnostico = alerta.get("alerta", "roas_bajo").replace("_", " ").title()
            accion = alerta.get("recomendacion") or (
                "Revisar foto/título" if "ctr" in diagnostico.lower() else (
                    "Revisar descripción/ficha" if "cvr" in diagnostico.lower() else "Evaluar pausa o cambio de tier"
                )
            )

            ws.cell(row=fila, column=1, value=grupo["item_ids"][0] if grupo["item_ids"] else "")
            ws.cell(row=fila, column=2, value=grupo.get("family_name", ""))
            ws.cell(row=fila, column=3, value=campania)
            ws.cell(row=fila, column=4, value=tier_actual.upper())
            ws.cell(row=fila, column=5, value=round(costo))
            ws.cell(row=fila, column=6, value=round(ingresos))
            ws.cell(row=fila, column=7, value=round(roas, 2))
            ws.cell(row=fila, column=8, value=f"{acos*100:.1f}%" if acos else "N/A")
            ws.cell(row=fila, column=9, value=diagnostico)
            ws.cell(row=fila, column=10, value=accion)
            ws.cell(row=fila, column=7).fill = PatternFill("solid", fgColor=_ROJO)
            fila += 1

        _autowidth(ws)

    # ------------------------------------------------------------------ #
    # Hoja 3: Alertas Urgentes
    # ------------------------------------------------------------------ #
    def _hoja_urgentes(self, wb, alertas_urgentes: list, semana: str) -> None:
        ws = wb.create_sheet("Alertas Urgentes")
        ws["A1"] = "⚠️ ALERTAS URGENTES - Productos con caída sostenida (3-4 días)"
        ws["A1"].font = Font(bold=True, size=12, color="FF0000")
        ws["A2"] = "Productos que venían con buen ROAS y empezaron a caer. Acción inmediata antes del lunes."

        _header(ws, 4, [
            "MLA", "Título", "Campaña",
            "ROAS prom. histórico", "ROAS últimos 3 días", "Caída (%)",
            "Días cayendo", "ACOS actual", "Posible causa", "Acción urgente",
        ])

        fila = 5
        for urgente in alertas_urgentes:
            roas_ult = urgente.get("roas_ultimos_dias", [])
            roas_ult_str = " / ".join(f"{r:.2f}" for r in roas_ult)
            acos = urgente.get("acos_actual")
            ws.cell(row=fila, column=1, value=urgente["item_ids"][0] if urgente["item_ids"] else "")
            ws.cell(row=fila, column=2, value=urgente.get("family_name", ""))
            ws.cell(row=fila, column=3, value=urgente.get("campania", ""))
            ws.cell(row=fila, column=4, value=urgente.get("roas_historico_prom"))
            ws.cell(row=fila, column=5, value=roas_ult_str)
            ws.cell(row=fila, column=6, value=f"{urgente.get('caida_pct', 0):.1f}%")
            ws.cell(row=fila, column=7, value=urgente.get("dias_cayendo"))
            ws.cell(row=fila, column=8, value=f"{acos*100:.1f}%" if acos else "N/A")
            ws.cell(row=fila, column=9, value="Revisar stock / competencia / precio")
            ws.cell(row=fila, column=10, value="Mover a Plata si persiste 1-2 días más")
            for col in range(1, 11):
                ws.cell(row=fila, column=col).fill = PatternFill("solid", fgColor=_AMARILLO)
            fila += 1

        _autowidth(ws)

    # ------------------------------------------------------------------ #
    # Hoja 4: Sin Ads - Candidatas
    # ------------------------------------------------------------------ #
    def _hoja_sin_ads(self, wb, candidatas: dict, semana: str) -> None:
        ws = wb.create_sheet("Sin Ads - Candidatas")
        ws["A1"] = "PUBLICACIONES SIN PUBLICIDAD - Candidatas a ingresar a Ads"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = "Publicaciones activas sin campaña de Ads. Ventas orgánicas en 30 días: verificar en ML (requiere orders API)."

        _header(ws, 4, [
            "MLA", "Título publicación", "Precio ($)", "Tier según ticket",
            "Ventas 30 días (*)", "Stock actual", "Variantes disp.", "Envío gratis?",
            "Campaña recomendada", "Prioridad", "Motivo y recomendación del agente",
        ])

        # color de prioridad
        colores_prioridad = {"Alta": _VERDE, "Media": _AMARILLO, "Baja": _ROJO}

        fila = 5
        for family_id, grupo in candidatas.items():
            prioridad = grupo.get("prioridad", "Media")
            ws.cell(row=fila, column=1, value=grupo["item_ids"][0] if grupo["item_ids"] else "")
            ws.cell(row=fila, column=2, value=grupo.get("family_name", ""))
            precio = grupo.get("precio", 0)
            ws.cell(row=fila, column=3, value=f"${precio:,.0f}" if precio else "Ver ML")
            ws.cell(row=fila, column=4, value=grupo.get("ticket", "?").capitalize())
            ws.cell(row=fila, column=5, value="(*) Ver en ML")
            ws.cell(row=fila, column=6, value=grupo.get("stock_total", "N/D"))
            ws.cell(row=fila, column=7, value=grupo.get("variantes_disponibles", "N/D"))
            ws.cell(row=fila, column=8, value=grupo.get("envio_gratis", "N/D"))
            ws.cell(row=fila, column=9, value=grupo.get("campania_recomendada", "testeo"))
            ws.cell(row=fila, column=10, value=prioridad)
            ws.cell(row=fila, column=10).fill = PatternFill("solid", fgColor=colores_prioridad.get(prioridad, _GRIS))
            ws.cell(row=fila, column=11, value=grupo.get("motivo", ""))
            ws.cell(row=fila, column=11).alignment = Alignment(wrap_text=True)
            fila += 1

        ws.cell(row=fila + 1, column=1, value="(*) Las ventas orgánicas de los últimos 30 días requieren la orders API de ML — verificar manualmente en el panel de vendedor.")
        ws.cell(row=fila + 1, column=1).font = Font(italic=True, color="808080")
        _autowidth(ws)

    # ------------------------------------------------------------------ #
    # Hoja 5: Acciones Semana
    # ------------------------------------------------------------------ #
    def _hoja_acciones(self, wb, acciones: list, semana: str) -> None:
        ws = wb.create_sheet("Acciones Semana")
        ws["A1"] = "ACCIONES DE LA SEMANA - Para aprobar por Telegram"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = f"Semana: {semana}  |  León aprueba o rechaza antes de que el agente ejecute"

        _header(ws, 4, ["#", "MLA", "Publicación", "Acción", "Motivo", "Prioridad", "Estado", "Observaciones León"])

        tipos_accionables = {"mover_tier", "agregar_a_testeo", "agregar_a_promo", "pausar", "ajustar_presupuesto"}
        accionables = [a for a in acciones if a.get("tipo") in tipos_accionables]

        fila = 5
        for n, accion in enumerate(accionables, start=1):
            tipo = accion.get("tipo", "")
            nombre = accion.get("family_name", "")
            mla = accion["item_ids"][0] if accion.get("item_ids") else ""

            if tipo == "mover_tier":
                accion_str = f"{accion.get('campania_origen', accion.get('tier_origen', ''))} → {accion.get('campania_destino', accion.get('tier_destino', ''))}"
                motivo = f"ROAS {accion.get('roas_reciente', 0):.2f}"
                if accion.get("poco_stock"):
                    motivo += " (⚠️ poco stock)"
                prioridad = "ALTA" if accion.get("tier_destino") == "oro" else "MEDIA"
            elif tipo == "pausar":
                accion_str = "Pausar / Sacar de Ads"
                motivo = f"ROAS {accion.get('roas_reciente', 0):.2f} sostenido bajo objetivo"
                prioridad = "ALTA"
            elif tipo == "agregar_a_promo":
                accion_str = "Agregar a promo ML"
                motivo = "Poco stock"
                prioridad = "MEDIA"
            elif tipo == "agregar_a_testeo":
                accion_str = f"Agregar a {accion.get('campania', 'testeo')}"
                motivo = "Publicación nueva sin campaña"
                prioridad = "BAJA"
            else:
                accion_str = tipo
                motivo = ""
                prioridad = "MEDIA"

            ws.cell(row=fila, column=1, value=n)
            ws.cell(row=fila, column=2, value=mla)
            ws.cell(row=fila, column=3, value=nombre)
            ws.cell(row=fila, column=4, value=accion_str)
            ws.cell(row=fila, column=5, value=motivo)
            ws.cell(row=fila, column=6, value=prioridad)
            ws.cell(row=fila, column=7, value="⏳ Pendiente")
            ws.cell(row=fila, column=8, value="")

            if prioridad == "ALTA":
                for col in range(1, 9):
                    ws.cell(row=fila, column=col).fill = PatternFill("solid", fgColor=_AMARILLO)
            fila += 1

        _autowidth(ws)

    # ------------------------------------------------------------------ #
    # Hoja 6: Historial de Cambios
    # ------------------------------------------------------------------ #
    def _hoja_historial(self, wb, changes_history: list) -> None:
        ws = wb.create_sheet("Historial Cambios")
        ws["A1"] = "HISTORIAL ACUMULATIVO DE CAMBIOS - Shaffe MercadoLibre Ads"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = "Cada semana el agente agrega una fila por cada cambio ejecutado. Registro histórico para aprender."

        _header(ws, 4, ["Fecha", "MLA", "Publicación", "Cambio realizado", "ROAS antes", "ROAS después (7 días)", "Resultado", "Tier origen", "Tier destino"])

        for fila, cambio in enumerate(changes_history, start=5):
            ws.cell(row=fila, column=1, value=cambio.get("fecha", ""))
            ws.cell(row=fila, column=2, value=cambio.get("mla", ""))
            ws.cell(row=fila, column=3, value=cambio.get("publicacion", ""))
            ws.cell(row=fila, column=4, value=cambio.get("cambio", ""))
            ws.cell(row=fila, column=5, value=cambio.get("roas_antes"))
            ws.cell(row=fila, column=6, value=cambio.get("roas_despues_7d"))
            ws.cell(row=fila, column=7, value=cambio.get("resultado", "pendiente"))
            ws.cell(row=fila, column=8, value=cambio.get("tier_origen", ""))
            ws.cell(row=fila, column=9, value=cambio.get("tier_destino", ""))

        _autowidth(ws)
