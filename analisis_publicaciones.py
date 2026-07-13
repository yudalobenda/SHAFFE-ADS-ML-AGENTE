"""Analisis completo de publicaciones SHAFFE para decision de ads.

Genera un Excel con 3 hojas:
  1. Sin Ads - Checklist: publicaciones fuera de ads con semaforo de preparacion
  2. En Ads - Performance: las que ya estan corriendo con sus metricas
  3. Resumen: totales y distribucion

Ejecutar con: python analisis_publicaciones.py
"""
from __future__ import annotations

import json
import os
import sys
from datetime import date, timedelta

from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(__file__)
load_dotenv(dotenv_path=os.path.join(BASE_DIR, ".env"))

from core.ml_client import MLClient, MLClientError
import core.campaign_rules as reglas

MEMORY_DIR = os.path.join(BASE_DIR, "memory")
LOGS_DIR = os.path.join(BASE_DIR, "logs")

VERDE    = "C6EFCE"
AMARILLO = "FFEB9C"
ROJO     = "FFC7CE"
GRIS     = "D9D9D9"
AZUL     = "1F4E79"

LISTING_LABELS = {
    "gold_pro":     "Gold Pro",
    "gold_special": "Gold Special",
    "silver":       "Clasica",
    "bronze":       "Bronce",
    "free":         "Gratuita",
}

LISTING_OK = {"gold_special", "gold_pro"}


def _header(ws, fila, cols, color=AZUL):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF" if color == AZUL else "000000")
    for i, v in enumerate(cols, 1):
        c = ws.cell(row=fila, column=i, value=v)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def _autowidth(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 55)


def _listing_label(tipo):
    return LISTING_LABELS.get(tipo, tipo or "Desconocido")


def _checklist(item, en_ads=False):
    """Evalua un item contra el checklist pre-ads. Devuelve lista de checks."""
    checks = []

    listing = item.get("listing_type_id", "")
    if listing in LISTING_OK:
        checks.append(("Tipo listing", True, _listing_label(listing)))
    else:
        checks.append(("Tipo listing", False, f"{_listing_label(listing)} — cambiar a Gold Special"))

    stock = item.get("available_quantity", 0)
    if stock >= 10:
        checks.append(("Stock", True, f"{stock} uds"))
    elif stock >= 5:
        checks.append(("Stock", None, f"{stock} uds — justo, reponer antes de escalar"))
    else:
        checks.append(("Stock", False, f"{stock} uds — insuficiente, reponer primero"))

    envio = (item.get("shipping") or {}).get("free_shipping", False)
    if envio:
        checks.append(("Envio gratis", True, "Si"))
    else:
        checks.append(("Envio gratis", False, "No — activar envio gratis mejora conversion"))

    precio = item.get("price", 0)
    if precio and precio > 0:
        checks.append(("Precio", True, f"${precio:,.0f}"))
    else:
        checks.append(("Precio", None, "Verificar en ML"))

    sold = item.get("sold_quantity", 0)
    if sold >= 20:
        checks.append(("Ventas historicas", True, f"{sold} vendidas — demanda probada"))
    elif sold >= 5:
        checks.append(("Ventas historicas", None, f"{sold} vendidas — track record incipiente"))
    else:
        checks.append(("Ventas historicas", False, f"{sold} vendidas — sin historico, mas riesgo"))

    return checks


def _score(checks):
    """0-100 basado en checks: True=pasa, None=parcial, False=no pasa."""
    pesos = {"Tipo listing": 30, "Stock": 25, "Envio gratis": 25, "Precio": 5, "Ventas historicas": 15}
    total = 0
    for nombre, estado, _ in checks:
        peso = pesos.get(nombre, 10)
        if estado is True:
            total += peso
        elif estado is None:
            total += peso * 0.5
    return round(total)


def _semaforo(score):
    if score >= 75:
        return "LISTA"
    if score >= 45:
        return "CASI LISTA"
    return "NECESITA ARREGLO"


def _semaforo_color(label):
    return {"LISTA": VERDE, "CASI LISTA": AMARILLO, "NECESITA ARREGLO": ROJO}.get(label, GRIS)


def _fix_list(checks):
    """Devuelve string con los puntos que fallaron."""
    items = [f"{n}: {d}" for n, ok, d in checks if ok is False]
    return " | ".join(items) if items else "—"


def main():
    print("Cargando tokens ML...")
    with open(os.path.join(MEMORY_DIR, "ml_tokens.json"), encoding="utf-8") as f:
        tokens = json.load(f)
    campaign_ids_data = json.load(open(os.path.join(MEMORY_DIR, "campaign_ids.json"), encoding="utf-8"))

    ml = MLClient(tokens)

    advertiser_id = str(campaign_ids_data["advertiser_id"])
    site_id = campaign_ids_data["site_id"]
    campanias = campaign_ids_data["campanas"] if "campanas" in campaign_ids_data else campaign_ids_data.get("campañas", {})
    ids_campanas = {str(v): k for k, v in campanias.items() if v}

    # ── 1. Traer todos los items activos ────────────────────────────────────
    print("Trayendo items activos de ML...")
    try:
        resultado = ml.get_seller_items(status="active")
        item_ids_todos = resultado.get("results", [])
    except MLClientError as e:
        print(f"Error al traer items: {e}")
        sys.exit(1)

    print(f"  {len(item_ids_todos)} items activos encontrados.")

    # ── 2. Traer detalle de cada item (en lotes de 20) ───────────────────────
    print("Trayendo detalle de items (listing_type, stock, precio, envio, ventas)...")
    items_detalle = ml.get_items_multiget(
        item_ids_todos[:200],
        attributes="id,title,listing_type_id,available_quantity,price,shipping,sold_quantity,status,category_id",
    )
    items_por_id = {item["id"]: item for item in items_detalle}
    print(f"  {len(items_por_id)} items con detalle.")

    # ── 3. Traer ads activos para saber cuales estan en campanas ─────────────
    print("Trayendo ads activos en campanas (ultimos 30 dias)...")
    date_to = date.today().isoformat()
    date_from = (date.today() - timedelta(days=30)).isoformat()
    try:
        ads = ml.search_ads_todas(site_id, advertiser_id, date_from, date_to)
    except MLClientError as e:
        print(f"Error al traer ads: {e}")
        ads = []

    # Construir mapa item_id → campana + metricas
    en_ads: dict = {}
    for ad in ads:
        campana_id = str(ad.get("campaign_id", ""))
        campana_nombre = ids_campanas.get(campana_id, campana_id)
        iid = ad.get("item_id", "")
        if not iid:
            continue
        metricas = ad.get("metrics") or {}
        cost = metricas.get("cost", 0.0)
        direct = metricas.get("direct_amount", 0.0)
        indirect = metricas.get("indirect_amount", 0.0)
        ingresos = direct + indirect
        roas = (ingresos / cost) if cost > 0 else 0.0
        en_ads[iid] = {
            "campana": campana_nombre,
            "costo": cost,
            "ingresos": ingresos,
            "roas": roas,
            "clicks": metricas.get("clicks", 0),
            "impresiones": metricas.get("prints", 0),
            "acos": metricas.get("acos"),
        }

    print(f"  {len(en_ads)} items encontrados en campanas.")

    # ── 4. Separar en ads / sin ads ──────────────────────────────────────────
    sin_ads = []
    con_ads = []
    for item_id, item in items_por_id.items():
        if item_id in en_ads:
            con_ads.append((item_id, item, en_ads[item_id]))
        else:
            sin_ads.append((item_id, item))

    print(f"  Sin ads: {len(sin_ads)} | Con ads: {len(con_ads)}")

    # ── 5. Generar Excel ─────────────────────────────────────────────────────
    print("Generando Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Hoja 1: Sin Ads ──────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Sin Ads - Checklist")
    ws1["A1"] = "PUBLICACIONES SIN ADS — Checklist de preparacion"
    ws1["A1"].font = Font(bold=True, size=13)
    ws1["A2"] = f"Generado: {date.today().isoformat()}  |  {len(sin_ads)} publicaciones sin campana de ads"
    ws1["A3"] = "Score 75-100: LISTA para ads | 45-74: CASI LISTA (arreglo menor) | 0-44: NECESITA ARREGLO antes de gastar"
    ws1["A3"].font = Font(italic=True, color="666666")

    cols_sin = [
        "MLA", "Titulo", "Tipo Listing", "Precio ($)", "Stock",
        "Envio gratis", "Ventas hist.", "Score (0-100)", "Estado",
        "Campana recomendada", "Que arreglar antes",
    ]
    _header(ws1, 5, cols_sin)

    # Ordenar por score desc
    filas_sin = []
    for item_id, item in sin_ads:
        checks = _checklist(item)
        score = _score(checks)
        estado = _semaforo(score)
        precio = item.get("price", 0)
        ticket = reglas.clasificar_ticket(precio) if precio else "medio"
        campana_rec = f"testeo_{ticket}"
        fix = _fix_list(checks)
        filas_sin.append((score, estado, item_id, item, checks, campana_rec, fix))

    filas_sin.sort(key=lambda x: -x[0])

    for fila_n, (score, estado, item_id, item, checks, campana_rec, fix) in enumerate(filas_sin, start=6):
        listing = item.get("listing_type_id", "")
        precio = item.get("price", 0)
        stock = item.get("available_quantity", 0)
        envio = "Si" if (item.get("shipping") or {}).get("free_shipping") else "No"
        sold = item.get("sold_quantity", 0)

        ws1.cell(row=fila_n, column=1, value=item_id)
        ws1.cell(row=fila_n, column=2, value=item.get("title", ""))
        ws1.cell(row=fila_n, column=3, value=_listing_label(listing))
        ws1.cell(row=fila_n, column=4, value=f"${precio:,.0f}" if precio else "?")
        ws1.cell(row=fila_n, column=5, value=stock)
        ws1.cell(row=fila_n, column=6, value=envio)
        ws1.cell(row=fila_n, column=7, value=sold)
        ws1.cell(row=fila_n, column=8, value=score)
        ws1.cell(row=fila_n, column=9, value=estado)
        ws1.cell(row=fila_n, column=10, value=campana_rec)
        ws1.cell(row=fila_n, column=11, value=fix)

        color = _semaforo_color(estado)
        for col in range(1, 12):
            ws1.cell(row=fila_n, column=col).fill = PatternFill("solid", fgColor=color)
        ws1.cell(row=fila_n, column=11).alignment = Alignment(wrap_text=True)

    # Colores en columnas de checklist individuales
    for fila_n, (score, estado, item_id, item, checks, _, _) in enumerate(filas_sin, start=6):
        for checks_n, (nombre, ok, desc) in enumerate(checks):
            col_extra = 12 + checks_n
            ws1.cell(row=fila_n, column=col_extra, value=desc)
            if ok is True:
                ws1.cell(row=fila_n, column=col_extra).fill = PatternFill("solid", fgColor=VERDE)
            elif ok is False:
                ws1.cell(row=fila_n, column=col_extra).fill = PatternFill("solid", fgColor=ROJO)
            else:
                ws1.cell(row=fila_n, column=col_extra).fill = PatternFill("solid", fgColor=AMARILLO)

    if filas_sin:
        checks_names = [c[0] for c in filas_sin[0][4]]
        for i, nombre in enumerate(checks_names):
            ws1.cell(row=5, column=12 + i, value=nombre).fill = PatternFill("solid", fgColor=AZUL)
            ws1.cell(row=5, column=12 + i).font = Font(bold=True, color="FFFFFF")

    _autowidth(ws1)

    # ── Hoja 2: En Ads ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("En Ads - Performance")
    ws2["A1"] = "PUBLICACIONES EN ADS — Performance ultimos 30 dias"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A2"] = f"Generado: {date.today().isoformat()}  |  {len(con_ads)} publicaciones activas en campanas"

    cols_ads = [
        "MLA", "Titulo", "Campana", "Tipo Listing", "Precio ($)", "Stock",
        "Envio gratis", "ROAS (30d)", "ACOS", "Inversion ($)", "Ingresos ($)",
        "Clicks", "Estado ROAS", "Accion recomendada",
    ]
    _header(ws2, 4, cols_ads)

    # Ordenar por ROAS desc
    con_ads_sorted = sorted(con_ads, key=lambda x: -x[2].get("roas", 0))

    for fila_n, (item_id, item, ads_data) in enumerate(con_ads_sorted, start=5):
        campana = ads_data["campana"]
        roas = ads_data["roas"]
        roas_obj = reglas.roas_target_campania(campana) if campana else 4.0
        listing = item.get("listing_type_id", "")
        precio = item.get("price", 0)
        stock = item.get("available_quantity", 0)
        envio = "Si" if (item.get("shipping") or {}).get("free_shipping") else "No"
        acos = ads_data.get("acos")

        if roas >= roas_obj:
            estado_roas = "OK"
            accion = "Mantener. Si ROAS > 6.5 sostenido 7 dias, evaluar subir a Oro."
            color = VERDE
        elif roas >= roas_obj * 0.7:
            estado_roas = "Revisar"
            accion = "Cerca del objetivo. Revisar foto principal si CTR es bajo."
            color = AMARILLO
        else:
            if roas == 0:
                estado_roas = "Sin datos"
                accion = "Sin conversiones. Verificar que ML le asigno presupuesto real (ver gasto)."
                color = GRIS
            else:
                estado_roas = "Bajo"
                accion = "ROAS bajo objetivo. Revisar foto/precio. Si no mejora en 7 dias, sacar de ads."
                color = ROJO

        ws2.cell(row=fila_n, column=1, value=item_id)
        ws2.cell(row=fila_n, column=2, value=item.get("title", ""))
        ws2.cell(row=fila_n, column=3, value=campana)
        ws2.cell(row=fila_n, column=4, value=_listing_label(listing))
        ws2.cell(row=fila_n, column=5, value=f"${precio:,.0f}" if precio else "?")
        ws2.cell(row=fila_n, column=6, value=stock)
        ws2.cell(row=fila_n, column=7, value=envio)
        ws2.cell(row=fila_n, column=8, value=round(roas, 2))
        ws2.cell(row=fila_n, column=9, value=f"{acos*100:.1f}%" if acos else "N/A")
        ws2.cell(row=fila_n, column=10, value=round(ads_data["costo"]))
        ws2.cell(row=fila_n, column=11, value=round(ads_data["ingresos"]))
        ws2.cell(row=fila_n, column=12, value=ads_data["clicks"])
        ws2.cell(row=fila_n, column=13, value=estado_roas)
        ws2.cell(row=fila_n, column=14, value=accion)

        for col in range(1, 15):
            ws2.cell(row=fila_n, column=col).fill = PatternFill("solid", fgColor=color)
        ws2.cell(row=fila_n, column=14).alignment = Alignment(wrap_text=True)

    _autowidth(ws2)

    # ── Hoja 3: Resumen ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Resumen")
    ws3["A1"] = "RESUMEN GENERAL"
    ws3["A1"].font = Font(bold=True, size=13)
    ws3["A2"] = f"Generado: {date.today().isoformat()}"

    listas = sum(1 for s, *_ in filas_sin if s >= 75)
    casi   = sum(1 for s, *_ in filas_sin if 45 <= s < 75)
    arreglo = sum(1 for s, *_ in filas_sin if s < 45)

    roas_ok  = sum(1 for _, item, ads_data in con_ads if ads_data["roas"] >= reglas.roas_target_campania(ads_data["campana"]))
    roas_mal = len(con_ads) - roas_ok

    listing_sin_gold = sum(1 for _, item in sin_ads if item.get("listing_type_id") not in LISTING_OK)
    listing_sin_gold_ads = sum(1 for _, item, _ in con_ads if item.get("listing_type_id") not in LISTING_OK)
    sin_envio_gratis = sum(1 for _, item in sin_ads if not (item.get("shipping") or {}).get("free_shipping"))

    filas_resumen = [
        ("", "", ""),
        ("PUBLICACIONES SIN ADS", "", ""),
        ("Total sin ads", len(sin_ads), ""),
        ("  Listas para entrar (score >= 75)", listas, "Verde en la hoja Sin Ads"),
        ("  Casi listas (score 45-74)", casi, "Amarillo — arreglo menor"),
        ("  Necesitan arreglo (score < 45)", arreglo, "Rojo — no gastar hasta arreglar"),
        ("  Sin Gold Special/Pro", listing_sin_gold, "Cambiar tipo listing antes de ads"),
        ("  Sin envio gratis", sin_envio_gratis, "Activar envio gratis mejora conversion"),
        ("", "", ""),
        ("PUBLICACIONES EN ADS", "", ""),
        ("Total en ads", len(con_ads), ""),
        ("  Con ROAS OK", roas_ok, "En objetivo o por encima"),
        ("  Con ROAS bajo objetivo", roas_mal, "Ver hoja En Ads para acciones"),
        ("  En ads sin Gold Special", listing_sin_gold_ads, "Pueden estar limitando el ROAS"),
    ]

    _header(ws3, 4, ["Metrica", "Valor", "Nota"])
    for fila_n, (metrica, valor, nota) in enumerate(filas_resumen, start=5):
        ws3.cell(row=fila_n, column=1, value=metrica)
        ws3.cell(row=fila_n, column=2, value=valor)
        ws3.cell(row=fila_n, column=3, value=nota)
        if "PUBLICACIONES" in str(metrica):
            ws3.cell(row=fila_n, column=1).font = Font(bold=True)

    _autowidth(ws3)

    # ── Guardar ──────────────────────────────────────────────────────────────
    os.makedirs(LOGS_DIR, exist_ok=True)
    nombre = f"analisis_publicaciones_{date.today().isoformat()}.xlsx"
    ruta = os.path.join(LOGS_DIR, nombre)
    wb.save(ruta)

    # Guardar tokens por si hubo refresh
    with open(os.path.join(MEMORY_DIR, "ml_tokens.json"), "w", encoding="utf-8") as f:
        json.dump(ml.tokens_actuales(), f, ensure_ascii=False, indent=2)

    print(f"\nExcel guardado en: {ruta}")
    print(f"\nResumen:")
    print(f"  Sin ads: {len(sin_ads)} publicaciones")
    print(f"    Listas: {listas} | Casi listas: {casi} | Necesitan arreglo: {arreglo}")
    print(f"  En ads: {len(con_ads)} publicaciones")
    print(f"    ROAS OK: {roas_ok} | ROAS bajo: {roas_mal}")


if __name__ == "__main__":
    main()
